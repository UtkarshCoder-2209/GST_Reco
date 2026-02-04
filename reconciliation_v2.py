import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import sys
import glob
import re
from collections import defaultdict
import datetime
import bisect
import gc

# --- Configuration ---
DEFAULT_TOLERANCE = 1.0  # Default amount tolerance if not specified

# --- Helper Functions ---

def classify_sheet(name):
    n = name.lower()
    if any(x in n for x in ['book', 'purchase', 'ledger', 'pr', 'register']):
        return 'BOOKS'
    if '2b' in n:
        return '2B'
    if '3b' in n:
        return '3B'
    return 'UNKNOWN'


def find_excel_files():
    patterns = ['*.xlsx', '*.xlsm', '*.xls']
    files = []
    for p in patterns:
        files.extend(glob.glob(p))
    return [f for f in files if not os.path.basename(f).startswith('~$')]

def normalize_header(h):
    if h is None: return ""
    return re.sub(r"\s+", " ", str(h).strip().lower())

def to_float(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    s = str(val).strip()
    if not s: return 0.0
    # specific cleanup
    s = s.replace(",", "").replace("\u00A0", "").replace("₹", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    # extract first float-like pattern
    match = re.search(r"-?\d+(\.\d+)?", s)
    if match:
        try:
            return float(match.group(0))
        except:
            return 0.0
    return 0.0

def get_header_map(ws):
    # Find header row (assume row 1 usually, or scan first few)
    # The previous script scanned row 1. Let's do that for simplicity.
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        norm = normalize_header(val)
        if norm:
            headers[norm] = col
    
    # Map essential columns
    mapping = {}
    
    def find_col(keywords):
        for k in keywords:
            for h in headers:
                if k in h:
                    return headers[h]
        return None

    mapping['gstin'] = find_col(['gstin', 'gst number', 'gst no', 'tin'])
    mapping['party'] = find_col(['party', 'name', 'legal name', 'trade name', 'supplier', 'customer'])
    mapping['igst'] = find_col(['igst', 'integrated tax'])
    mapping['cgst'] = find_col(['cgst', 'central tax'])
    mapping['sgst'] = find_col(['sgst', 'state tax', 'utgst'])
    mapping['inv_no'] = find_col(['invoice number', 'inv no', 'bill no'])
    mapping['date'] = find_col(['invoice date', 'date', 'inv dt'])
    
    # Fallback for tax if specific columns not found (sometimes just 'Tax Amount') - but user was specific about heads.
    # Assuming standard format based on previous script.
    
    return mapping

class Invoice:
    __slots__ = ['row_idx', 'gstin', 'igst', 'cgst', 'sgst', 'party', 'head', 'amount', 'source', 'match_id', 'match_type', 'matched_with_row']

    def __init__(self, row_idx, data, source_sheet):
        self.row_idx = row_idx
        self.gstin = str(data.get('gstin', '') or '').strip().upper()
        # Clean GSTIN: sometimes it might be empty or 'None' string
        if self.gstin in ['NONE', 'NAN']: self.gstin = ''
            
        self.igst = to_float(data.get('igst', 0))
        self.cgst = to_float(data.get('cgst', 0))
        self.sgst = to_float(data.get('sgst', 0))
        self.party = str(data.get('party', '') or '').strip()
        
        # Determine Head and Main Amount
        # USE LOWER THRESHOLD (0.01) to catch small differences
        if abs(self.igst) > 0.01:
            self.head = 'IGST'
            self.amount = self.igst
        elif abs(self.cgst) > 0.01 or abs(self.sgst) > 0.01:
            self.head = 'CGST/SGST'
            # Use CGST as the match amount.
            self.amount = self.cgst if abs(self.cgst) > 0.01 else self.sgst
        else:
            self.head = 'ZERO'
            self.amount = 0.0
            
        self.source = source_sheet
        self.match_id = None
        self.match_type = None # 'Perfect', 'Missing GSTIN', 'Amount Only'
        self.matched_with_row = None

    def __repr__(self):
        return f"Row:{self.row_idx} GST:{self.gstin} {self.head}:{self.amount}"

def read_invoices(ws, mapping, sheet_name):
    invoices = []
    # Data starts from row 2
    for r in range(2, ws.max_row + 1):
        data = {}
        # Check if row is empty
        is_empty = True
        for k, col_idx in mapping.items():
            if col_idx:
                val = ws.cell(row=r, column=col_idx).value
                data[k] = val
                if val is not None and str(val).strip() != "":
                    is_empty = False
            else:
                data[k] = None
                
        if not is_empty:
            inv = Invoice(r, data, sheet_name)
            # Include everything that is not strictly identified as ZERO tax
            # This allows Negative amounts (Returns) to be included.
            if inv.head != 'ZERO':
                invoices.append(inv)
    return invoices

def match_invoices(list1, list2, tolerance, name1, name2):
    # list1: invoices from sheet 1 (name1)
    # list2: invoices from sheet 2 (name2)
    
    match_counter = 1
    
    # Helper to check match
    def amounts_match(a, b, tol):
        return abs(a - b) <= tol

    # --- Phase 1: Perfect Match (Same GSTIN, Same Head, Same Amount) ---
    print("Running Phase 1: Perfect Matches...")
    
    # Index list2 by (GSTIN, Head) for faster lookup
    # Since we can have multiple matches, we use a list
    map2 = defaultdict(list)
    for inv in list2:
        if not inv.match_id:
            key = (inv.gstin, inv.head)
            map2[key].append(inv)
            
    for inv1 in list1:
        if inv1.match_id: continue
        
        key = (inv1.gstin, inv1.head)
        candidates = map2.get(key, [])
        
        best_match = None
        best_diff = float('inf')
        
        # Find best amount match within tolerance
        for candidate in candidates:
            if candidate.match_id: continue
            
            diff = abs(inv1.amount - candidate.amount)
            if diff <= tolerance and diff < best_diff:
                best_diff = diff
                best_match = candidate
        
        if best_match:
            # Link them
            inv1.match_id = match_counter
            best_match.match_id = match_counter
            inv1.match_type = "Perfect Match"
            best_match.match_type = "Perfect Match"
            inv1.matched_with_row = best_match.row_idx
            best_match.matched_with_row = inv1.row_idx
            match_counter += 1

    # --- Phase 2: Missing GSTIN (One side has GSTIN, other is blank, Head+Amount Match) ---
    # User might mean: If GSTIN is missing in one sheet but present in another.
    print("Running Phase 2: Missing GSTIN Matches...")
    
    # Re-index remaining unmatched items in list2 by Head only
    # But specifically target those with EMPTY GSTIN in list2 if list1 has GSTIN, or vice-versa
    
    # Sub-case A: List1 has GSTIN, List2 has NO GSTIN
    map2_no_gstin = defaultdict(list)
    for inv in list2:
        if not inv.match_id and not inv.gstin:
            map2_no_gstin[inv.head].append(inv)
            
    for inv1 in list1:
        if inv1.match_id: continue
        if not inv1.gstin: continue # We need GSTIN here to call it "Missing GSTIN match" (one side has it)
        
        candidates = map2_no_gstin.get(inv1.head, [])
        best_match = None
        best_diff = float('inf')
        
        for candidate in candidates:
            if candidate.match_id: continue
            diff = abs(inv1.amount - candidate.amount)
            if diff <= tolerance and diff < best_diff:
                best_diff = diff
                best_match = candidate
        
        if best_match:
            inv1.match_id = match_counter
            best_match.match_id = match_counter
            inv1.match_type = f"Match (Missing GSTIN in {name2})"
            best_match.match_type = f"Match (Missing GSTIN in {name2})"
            inv1.matched_with_row = best_match.row_idx
            best_match.matched_with_row = inv1.row_idx
            match_counter += 1
            
    # Sub-case B: List1 has NO GSTIN, List2 has GSTIN
    # Index list2 items that HAVE GSTIN
    map2_with_gstin = defaultdict(list)
    for inv in list2:
        if not inv.match_id and inv.gstin:
            map2_with_gstin[inv.head].append(inv)
            
    for inv1 in list1:
        if inv1.match_id: continue
        if inv1.gstin: continue 
        
        candidates = map2_with_gstin.get(inv1.head, [])
        best_match = None
        best_diff = float('inf')
        
        for candidate in candidates:
            if candidate.match_id: continue
            diff = abs(inv1.amount - candidate.amount)
            if diff <= tolerance and diff < best_diff:
                best_diff = diff
                best_match = candidate
                
        if best_match:
            inv1.match_id = match_counter
            best_match.match_id = match_counter
            inv1.match_type = f"Match (Missing GSTIN in {name1})"
            best_match.match_type = f"Match (Missing GSTIN in {name1})"
            inv1.matched_with_row = best_match.row_idx
            best_match.matched_with_row = inv1.row_idx
            match_counter += 1

    # Free up memory explicitly
    map2_no_gstin.clear()
    map2_with_gstin.clear()
    map2.clear()
    gc.collect()

    # --- Phase 3: Amount Match (GSTIN Mismatch / Both have different GSTINs or both missing) ---
    print("Running Phase 3: Amount Only Matches...")
    
    # Index remaining list2 by Head
    map2_any = defaultdict(list)
    for inv in list2:
        if not inv.match_id:
            map2_any[inv.head].append(inv)
            
    # Sort candidates by amount for binary search
    sorted_candidates_map = {}
    candidates_amounts_map = {} # Pre-calculate amounts for bisect
    
    for head, items in map2_any.items():
        # filtering out items that might have been matched in previous phases
        valid_items = [x for x in items if not x.match_id]
        valid_items.sort(key=lambda x: x.amount)
        sorted_candidates_map[head] = valid_items
        candidates_amounts_map[head] = [x.amount for x in valid_items]
        
    for inv1 in list1:
        if inv1.match_id: continue
        
        candidates = sorted_candidates_map.get(inv1.head, [])
        candidate_amounts = candidates_amounts_map.get(inv1.head, [])
        
        if not candidates: continue
        
        # Binary search using pre-calculated list
        idx = bisect.bisect_left(candidate_amounts, inv1.amount)
        
        best_match = None
        best_diff = float('inf')
        
        # Check right side (including idx)
        for i in range(idx, len(candidates)):
            # Optimization: strictly sorted
            diff = abs(candidates[i].amount - inv1.amount)
            if diff > tolerance:
                break 
            
            if candidates[i].match_id: continue
            
            if diff < best_diff:
                best_diff = diff
                best_match = candidates[i]
                
        # Check left side
        for i in range(idx - 1, -1, -1):
            diff = abs(candidates[i].amount - inv1.amount)
            if diff > tolerance:
                break
                
            if candidates[i].match_id: continue
            
            if diff < best_diff:
                best_diff = diff
                best_match = candidates[i]
        
        if best_match:
            inv1.match_id = match_counter
            best_match.match_id = match_counter
            
            # Check if GSTINs are present but differ
            if inv1.gstin and best_match.gstin and inv1.gstin != best_match.gstin:
                match_desc = "Probable Match (GSTIN Mismatch)"
            else:
                 # Both empty?
                match_desc = "Match (No GSTINs)"
                
            inv1.match_type = match_desc
            best_match.match_type = match_desc
            inv1.matched_with_row = best_match.row_idx
            best_match.matched_with_row = inv1.row_idx
            match_counter += 1

    return match_counter

def write_results(ws, invoices, this_name, other_name):
    # Determine max column to append
    max_col = ws.max_column
    
    # Headers
    ws.cell(row=1, column=max_col+1, value="Match_Status")
    ws.cell(row=1, column=max_col+2, value="Match_ID")
    ws.cell(row=1, column=max_col+3, value="Matched_Row_Idx")
    
    # Map by row index for easy writing
    inv_map = {inv.row_idx: inv for inv in invoices}
    
    for r in range(2, ws.max_row + 1):
        if r in inv_map:
            inv = inv_map[r]
            # Descriptive status
            status = inv.match_type if inv.match_type else f"Missing in {other_name}"
            
            ws.cell(row=r, column=max_col+1, value=status)
            ws.cell(row=r, column=max_col+2, value=inv.match_id)
            ws.cell(row=r, column=max_col+3, value=inv.matched_with_row)
        else:
            # Rows that were skipped (e.g. zero value)
            ws.cell(row=r, column=max_col+1, value="Ignored/Zero")

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", help="Path to Excel file")
    parser.add_argument("--sheet1", help="Name of first sheet")
    parser.add_argument("--sheet2", help="Name of second sheet")
    parser.add_argument("--tolerance", type=float, default=DEFAULT_TOLERANCE, help="Amount tolerance")
    args = parser.parse_args()

    files = find_excel_files()
    path = None
    
    if args.file:
        path = args.file
    elif len(files) == 1:
        path = files[0]
        print(f"Auto-selected file: {path}")
    elif not files:
        print("No Excel files found.")
        return
    else:
        print("Found files:")
        for i, f in enumerate(files):
            print(f"{i+1}: {f}")
        choice = input("Select file (number): ")
        try:
            path = files[int(choice)-1]
        except:
            print("Invalid choice.")
            return

    tolerance = args.tolerance
    if not args.tolerance and not args.file: # If interactive mode
        tol_input = input(f"Enter amount tolerance (default {DEFAULT_TOLERANCE}): ")
        try:
             if tol_input.strip(): tolerance = float(tol_input)
        except:
            pass
            
    print(f"Loading {path}...")
    wb = load_workbook(path)
    sheets = wb.sheetnames
    
    name1 = args.sheet1
    name2 = args.sheet2
    
    # Auto-detection logic if names not provided
    if not name1 and not name2:
        classified = {s: classify_sheet(s) for s in sheets}
        
        books_sheets = [s for s, t in classified.items() if t == 'BOOKS']
        gstr2b_sheets = [s for s, t in classified.items() if t == '2B']
        gstr3b_sheets = [s for s, t in classified.items() if t == '3B']
        
        # Priority Logic:
        # 1. Books vs 2B
        # 2. Books vs 3B
        # 3. 3B vs 2B (Reconciling Claimed vs Available)
        
        if books_sheets and gstr2b_sheets:
            name1 = books_sheets[0]
            name2 = gstr2b_sheets[0]
        elif books_sheets and gstr3b_sheets:
            name1 = books_sheets[0]
            name2 = gstr3b_sheets[0]
        elif gstr3b_sheets and gstr2b_sheets:
            name1 = gstr3b_sheets[0]
            name2 = gstr2b_sheets[0]
        elif len(sheets) == 2:
            # Fallback: Just take the two available sheets
            name1 = sheets[0]
            name2 = sheets[1]
            
        if name1 and name2 and name1 != name2:
            print(f"Auto-detected sheets: {name1} (Base) vs {name2} (Counter)")

    if not name1 or not name2:
        if len(sheets) < 2:
            print("Workbook must have at least 2 sheets.")
            return
        
        print("Sheets available:")
        for i, s in enumerate(sheets):
            print(f"{i+1}: {s}")
            
        if not name1:
            s1_idx = int(input("Select Sheet 1 (Book/Base): ")) - 1
            name1 = sheets[s1_idx]
        if not name2:
            s2_idx = int(input("Select Sheet 2 (Counterparty/2B/3B): ")) - 1
            name2 = sheets[s2_idx]
    
    # Ensure they are distinct
    if name1 == name2:
        print("Error: Sheet 1 and Sheet 2 must be different.")
        return

    ws1 = wb[name1]
    ws2 = wb[name2]
    
    print("Analyzing headers...")
    map1 = get_header_map(ws1)
    map2 = get_header_map(ws2)
    
    # Verify mappings
    print(f"Sheet '{name1}' columns detected: { {k:v for k,v in map1.items() if v} }")
    print(f"Sheet '{name2}' columns detected: { {k:v for k,v in map2.items() if v} }")
    
    # Read data
    print("Reading data...")
    inv1 = read_invoices(ws1, map1, name1)
    inv2 = read_invoices(ws2, map2, name2)
    
    print(f"Loaded {len(inv1)} records from {name1}")
    print(f"Loaded {len(inv2)} records from {name2}")
    
    # Match
    count = match_invoices(inv1, inv2, tolerance, name1, name2)
    print(f"Total Matches Found: {count-1}")
    
    # Write Results
    print("Writing results...")
    write_results(ws1, inv1, name1, name2)
    write_results(ws2, inv2, name2, name1)
    
    # Save
    # Save
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = f"{os.path.splitext(path)[0]}_RECON_{ts}.xlsx"
    wb.save(out_file)
    print(f"Saved to {out_file}")
    return out_file

def process_reconciliation(path, sheet1=None, sheet2=None, tolerance=DEFAULT_TOLERANCE):
    print(f"Processing {path}...")
    wb = load_workbook(path)
    sheets = wb.sheetnames
    
    name1 = sheet1
    name2 = sheet2
    
    # Auto-detection logic if names not provided (Reused logic)
    if not name1 and not name2:
        classified = {s: classify_sheet(s) for s in sheets}
        books_sheets = [s for s, t in classified.items() if t == 'BOOKS']
        gstr2b_sheets = [s for s, t in classified.items() if t == '2B']
        gstr3b_sheets = [s for s, t in classified.items() if t == '3B']
        
        if books_sheets and gstr2b_sheets:
            name1 = books_sheets[0]
            name2 = gstr2b_sheets[0]
        elif books_sheets and gstr3b_sheets:
            name1 = books_sheets[0]
            name2 = gstr3b_sheets[0]
        elif gstr3b_sheets and gstr2b_sheets:
            name1 = gstr3b_sheets[0]
            name2 = gstr2b_sheets[0]
        elif len(sheets) == 2:
            name1 = sheets[0]
            name2 = sheets[1]
            
    # Validations
    if not name1 or not name2:
        raise ValueError(f"Could not auto-detect sheets. and none provided. Available: {sheets}")
        
    if name1 == name2:
        raise ValueError("Sheet 1 and Sheet 2 must be different.")
        
    if name1 not in sheets or name2 not in sheets:
        raise ValueError(f"One of the sheets not found. Available: {sheets}")

    ws1 = wb[name1]
    ws2 = wb[name2]
    
    map1 = get_header_map(ws1)
    map2 = get_header_map(ws2)
    
    # Validation logic
    required_cols = ['gstin']
    tax_cols = ['igst', 'cgst', 'sgst']
    
    def validate_map(m, sheet_name):
        missing = [c for c in required_cols if not m.get(c)]
        if missing:
            raise ValueError(f"Sheet '{sheet_name}' is missing required columns: {', '.join(missing)}")
        
        # Check if at least one tax column exists? Or just warn?
        # User requested specific error like "No IGST column found"
        # Let's be strict about tax columns if possible, or at least check if ALL are missing
        if not any(m.get(c) for c in tax_cols):
             raise ValueError(f"Sheet '{sheet_name}' has no tax columns (IGST, CGST, or SGST found).")

    validate_map(map1, name1)
    validate_map(map2, name2)
    
    inv1 = read_invoices(ws1, map1, name1)
    inv2 = read_invoices(ws2, map2, name2)
    
    match_invoices(inv1, inv2, tolerance, name1, name2)
    
    write_results(ws1, inv1, name1, name2)
    write_results(ws2, inv2, name2, name1)
    
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = f"{os.path.splitext(path)[0]}_RECON_{ts}.xlsx"
    wb.save(out_file)
    return out_file

if __name__ == "__main__":
    main()

